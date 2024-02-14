# -*- coding: utf-8 -*-
"""
Created on Tue Aug 22 10:40:40 2023

@author: Doumit
"""

import tkinter as tk
from tkinter import messagebox 
from openpyxl import load_workbook
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from datetime import datetime
import time
import win32com.client as win32
import os
import base64
import requests
import pyodbc

def cliente_equipamento(cliente):

    # Configuração da conexão com o banco de dados
    server = ''  # Substitua pelo nome do servidor SQL Server
    database = ''  # Substitua pelo nome do banco de dados
    username = ''  # Substitua pelo seu nome de usuário
    password = ''  # Substitua pela sua senha
    
    # String de conexão
    conn_str = f'DRIVER={{ODBC Driver 17 for SQL Server}};SERVER={server};DATABASE={database};UID={username};PWD={password}'
    
    
    # Estabeleça a conexão com o banco de dados
    conn = pyodbc.connect(conn_str)
    
    # Consulta SQL
    query = f"""
    select * from relatorio.vw_cliente_pacote
    where nm_cliente = '{cliente}'
    """ 
    
    # Execute a consulta
    cursor = conn.cursor()
    cursor.execute(query)

    # Recupere o valor da coluna "ds_equipamento" da primeira linha
    row = cursor.fetchone()
    
    if row is not None:
        equip = row.ds_equipamento
    else:
        equip = None

    # Feche a conexão com o banco de dados
    conn.close()

    return equip

# modulo =  modulo_servidor[0]
root = tk.Tk()
root.withdraw()

def last_chat_id(token):
    try:
        url = f"https://api.telegram.org/bot{token}/getUpdates"
        response = requests.get(url)
        if response.status_code == 200:
            json_msg = response.json()
            for json_result in reversed(json_msg['result']):
                message_keys = json_result['message'].keys()
                if ('new_chat_member' in message_keys) or ('group_chat_created' in message_keys):
                    return json_result['message']['chat']['id']
            print('Nenhum grupo encontrado')
        else:
            print('A resposta falhou, código de status:', response.status_code)
    except Exception as e:
        print("Erro no getUpdates:", e)

def send_message(token, chat_id, message):
    try:
        data = {"chat_id": chat_id, "text": message}
        url = f"https://api.telegram.org/bot{token}/sendMessage"
        response = requests.post(url, json=data)
        if response.status_code == 200:
            print("Mensagem enviada com sucesso!")
        else:
            print("Erro ao enviar a mensagem. Código de status:", response.status_code)
    except Exception as e:
        print("Erro no sendMessage:", e)
        
def send_document(token, chat_id, document_path):
    try:
        url = f"https://api.telegram.org/bot{token}/sendDocument"
        files = {'document': open(document_path, 'rb')}
        data = {'chat_id': chat_id}
        response = requests.post(url, data=data, files=files)
        if response.status_code == 200:
            print("Documento enviado com sucesso!")
        else:
            print("Erro ao enviar o documento. Código de status:", response.status_code)
    except Exception as e:
        print("Erro no sendDocument:", e)

token = ''


chat_id = ''



def enviar():
    try:
        diretorio_atual = os.getcwd()
        
        # Carrega o arquivo do Excel
        workbook = load_workbook('OS_VIVO.xlsm', data_only=True, keep_vba=True)
        
        # Seleciona a aba desejada
        worksheet = workbook['base']
        
        # Obtém o caminho completo para o arquivo Excel
        caminho_arquivo_excel = os.path.abspath('OS_VIVO.xlsm')
        
        # Abre o aplicativo Excel
        excel = win32.gencache.EnsureDispatch('Excel.Application')
        
        # Abre o arquivo Excel
        workbook = excel.Workbooks.Open(caminho_arquivo_excel)
        
        # Executa a macro VBA
        excel.Application.Run("OS_VIVO.xlsm!GerarOS")
        
        # Salva as alterações
        workbook.Save()
        
        # Fecha o arquivo Excel
        workbook.Close()
        
        # fecha
        excel.Quit()
        
        # Configurar as informações do seu e-mail e servidor SMTP
        remetente = ''
        senha = ''
        
        # Configurar o servidor SMTP do Outlook
        servidor_smtp = 'smtp.office365.com'
        porta_smtp = 587
        
        # Criar um dicionário para agrupar as placas com a mesma data, hora e endereço
        placas_agrupadas = {}
        
        # Iterar sobre as linhas do Excel (começando da segunda linha, presumindo que a primeira linha seja o cabeçalho)
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            
            if row[0] is None:
                break
            
            nome_cliente = row[2]
            
            try:
                # Chame a função que pode gerar um erro
                equipamento = cliente_equipamento(nome_cliente)  # Exemplo de erro, divisão por zero
            except Exception as e:
                messagebox.showerror("Erro", f"Ocorreu um erro:\n{e}")
            
            destinatario = row[26]  # Coluna do Excel com os endereços de e-mail dos destinatários
            data_hora = row[1]
            endereco = row[5]
            servico = row[3]
            responsavel = row[19]
            dados_tecnico_nome = row[21]
            dados_tecnico_doc = row[27]
            dados_tecnico_veic = row[28]
            dados_id_incidente = row[23]
            dados_placa = row[0]
            telefone = row[22]
            nu_os = row[4]
            cidade = row[8]
            uf = row[9]
        
        
            def obter_saudacao():
                hora_atual = datetime.now().hour
        
                if hora_atual < 12:
                    return 'Bom dia'
                elif hora_atual < 18:
                    return 'Boa tarde'
                else:
                    return 'Boa noite'
        
            # Exemplo de uso
            saudacao = obter_saudacao()
            
            # Verificar se já existe uma entrada no dicionário para essa data, hora e endereço
            chave = (data_hora, endereco,)
            if chave in placas_agrupadas:
                # Se já existir, adicione a placa atual ao valor (que é uma lista) da entrada do dicionário
                placas_agrupadas[chave].append((destinatario, equipamento, nome_cliente, data_hora, endereco, servico, responsavel, dados_tecnico_nome, dados_tecnico_doc, dados_tecnico_veic, dados_id_incidente, dados_placa, telefone, nu_os, cidade, uf))  # Modificado aqui para guardar placa e nu_os
            else:
                # Se não existir, crie uma nova entrada no dicionário com a placa atual e nu_os como o primeiro item da lista
                placas_agrupadas[chave] = [(destinatario, equipamento, nome_cliente, data_hora, endereco, servico, responsavel, dados_tecnico_nome, dados_tecnico_doc, dados_tecnico_veic, dados_id_incidente, dados_placa, telefone, nu_os, cidade, uf)]  # Modificado aqui para guardar placa e nu_os
        
        # Agora, vamos enviar as mensagens agrupadas no WhatsApp com os anexos corretos
        for chave, placas in placas_agrupadas.items():
            data_hora, endereco = chave
            
            # Obtém o caminho completo para a pasta OS_GERADAS
            caminho_os_geradas = os.path.join(diretorio_atual, 'OS_GERADAS')
            
            # Obter a lista de caminhos de arquivos de anexos correspondentes às placas agrupadas
            caminhos_arquivos_anexos = [os.path.join(caminho_os_geradas, f'{placa[11]} - {placa[13]}.pdf') for placa in placas]
        
            
            placas_servicos = "\n".join([f"{placa[11]}   {placa[5]}" for placa in placas])
            mensagem_telegram = f"""*Segue confirmação de agendamento:* 
            \n*Cliente:* VIVO 
            \n*Empresa:* {placas[0][2]} 
            \n*Data:* {data_hora} 
            \n*Endereço:* {endereco} 
            \n*Contato:* {', '.join(placa[6] for placa in placas)} 
            \n*Placas e Serviços*\n{placas_servicos}
            \n*Itens Instalados:* {placas[0][1]} """
           
            send_message(token, chat_id, mensagem_telegram)
            
            emoji_bola = u'\U00002b55'  
            mensagem_dois = f"""
            Segue ordem de serviço preenchida com as informações básicas do serviço.
            
            \nPeço por favor *IMPRIMIR e PREENCHER* : *DISPOSITIVO RETIRADO / INSTALADO* e *ASSINATURA  DO CLIENTE NOS DOIS ÚLTIMOS CAMPOS*       
            
            \n{emoji_bola}ANTES DE INICIAR QUALQUER SERVIÇO, DEVE SER GRAVADO UM VÍDEO DAS CONDIÇÕES DO VEÍCULO OU DO EQUIPAMENTO INSTALADO. EM CASOS DE MANUTENÇÃO OU RETIRADA, MOSTRAR EVIDÊNCIAS QUE PODEM TER HAVIDO MANIPULAÇÃO DO EQUIPAMENTO.
            
            \n{emoji_bola}O NÃO ENVIO DA FOTO DA ORDEM DE SERVIÇO  PODE IMPACTAR NO SEU PRÓXIMO CICLO DE PAGAMENTO                         

            \n{emoji_bola}EM CASO DE IMPRODUTIVA ENVIAR A ORDEM DE SERVIÇO ASSINADA PELO CLIENTE                      
            
            \n*EM CASO DE IMPRODUTIVA ENVIAR A ORDEM DE SERVIÇO ASSINADA PELO CLIENTE*
            """
            
            send_message(token, chat_id, mensagem_dois)
            
            # enviar anexos dentro da pasta OS_Geradas
            
            if len(caminhos_arquivos_anexos) == 1:
                send_document(token, chat_id, caminhos_arquivos_anexos[0])
                time.sleep(2)
            else:
                for caminho_arquivo_anexo in caminhos_arquivos_anexos:
                    send_document(token, chat_id, caminho_arquivo_anexo)
                    time.sleep(2)
            
            emoji_check = u'\U00002705'  
            mensagem_quatro = f"""
            *CHECK LISTA ATENDIMENTO VIVO*                         
            \nCheck List para realizar ao iniciar atendimento de cada veículo               
            \nFoto da placa{emoji_check}                      
            \nFoto do hodometro{emoji_check}                       
            \nFoto + Vídeo de como{emoji_check} encontrou o dispositivo{emoji_check} *(se há manipulação ou rastreador solto)*                        
            \nFoto do numero do dispositivo{emoji_check}                        
            \nFoto do numero do Identificador {emoji_check}                        
            \nFoto de como estão os leds{emoji_check}                        
            \nInformar se o veiculo esta bipando quando liga{emoji_check}                        
            \nInformar se o condutor possui a tag de identificação{emoji_check}                        
            \nInformar o nome do responsável ou motorista que esta acompanhando{emoji_check}                       
            \n*O RASTREADOR PRECISA ESTAR BEM FIXADO E INSTALADO NO PAINEL CENTRAL ATRAS DO TUDO DE AR*     
            """
            
            send_message(token, chat_id, mensagem_quatro)
            
            time.sleep(7)
            
            caminho_imagem = 'assinatura_vivo.png' 
        
            # Ler o conteúdo da imagem e codificar em base64
            with open(caminho_imagem, 'rb') as arquivo_imagem:
                conteudo_imagem_base64 = arquivo_imagem.read()
                conteudo_imagem_base64 = base64.b64encode(conteudo_imagem_base64).decode('utf-8')
            
            destinatarios = destinatario.split(';')
            
            # Criar uma mensagem de e-mail
            mensagem = MIMEMultipart()
            mensagem['From'] = remetente    
            mensagem['To'] = ", ".join(destinatarios)
            mensagem['Subject'] = f"""CONFIRMAÇÃO DE AGENDAMENTO - {placas[0][2]} - {placas[0][5]} - {placas[0][14]} /{placas[0][15]}"""
            
            corpo_email = f"""
            <html>
            
            <style>
            
            table {{
                border-collapse: collapse;
            }}
            
            th, td {{
                text-align: center;
                padding: 2px;
            }}
            
            td {{
                border: 1px solid #ddd; /* Bordas cinzas para as células */
            }}
            
            th {{
                background-color: #99CAFE;
                color: balck;
                border: 1px solid white;
            }}
            </style>

            <span>{saudacao}<br>
            Estimamos que esteja bem,</span><br><br>
        
            <h2>Conforme solicitado segue confirmação de agendamento:</h2>
        
            <span style="font-weight: bold">Data:</span><span> {data_hora}</span><br>
            <span style="font-weight: bold">Endereço:</span> {endereco}</span><br>
            <span style="font-weight: bold">Responsável:</span> {', '.join(placa[6] for placa in placas)}</span><br>
        
            <h3 style="font-weight: bold">Detalhes do Chamado:</h3>
            
            <table>
            <tr>
            <th>ID DO INCIDENTE</th>
            <th>PLACA</th>
            <th>SERVIÇO</th>
            
            </tr>
            {"".join(f"<tr><td>{placa[10]}</td><td>{placa[11]}</td><td>{placa[5]}</td></tr>" for placa in placas)}
            </table><br>
        
            <p>Atenção: Caso ocorra algum imprevisto havendo a necessidade de cancelar atendimento, para que não ocorra visita improdutiva realizar contato através do 10315 opção 2 cód. 1629, ou nesse e-mail até 24 horas úteis antes do atendimento.</p><br>
        
             """
        
            mensagem.attach(MIMEText(corpo_email, 'html'))
        
            # Estabelecer uma conexão segura com o servidor SMTP
            conexao_smtp = smtplib.SMTP(servidor_smtp, porta_smtp)
            conexao_smtp.starttls()
        
            # Autenticar com o servidor SMTP
            conexao_smtp.login(remetente, senha)
        
            # Enviar o e-mail
            conexao_smtp.sendmail(remetente, destinatarios, mensagem.as_string())
        
            # Encerrar a conexão com o servidor SMTP
            conexao_smtp.quit()

        # Exibe uma mensagem de sucesso
        messagebox.showinfo("Sucesso", "E-mails e mensagens no Telegram enviados com sucesso!")
        
    except Exception as e:
        # Exibe uma mensagem de erro
        print("Erro", f"Ocorreu um erro:\n{e}")
        messagebox.showerror("Erro", f"Ocorreu um erro:\n{e}")
        
enviar()